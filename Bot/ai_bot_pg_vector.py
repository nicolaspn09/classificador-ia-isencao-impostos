import os
import re
import sys
import json
from openpyxl import load_workbook
from typing import Optional, Any, Dict, List
from langchain_groq import ChatGroq
from langchain.chains.combine_documents import create_stuff_documents_chain # Retorna um chain pronto
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_core.prompts import ChatPromptTemplate, PromptTemplate
from dotenv import load_dotenv

# Funcões para o agente
from langchain.agents import create_react_agent, AgentExecutor
from langchain import tools

sys.path.append(r"C:\rpa\Python")
from Classes.Postgres.Postgres.ConectaPGVector import ConectaPGVector


# ATENÇÃO: Use o mesmo collection_name que você usou no script de gravação da planilha!
COLLECTION_NAME_PLANILHA = "isencoes_de_produtos"
pg_vector_connector = ConectaPGVector()

load_dotenv()
groq_api_key = os.getenv("GROQ_API_KEY")

model = ChatGroq(
    model = "meta-llama/llama-4-maverick-17b-128e-instruct",
)

# Cria uma conexão com o banco de dados de vetores
persist_directory = rf"C:\rpa\Python\Modelo Isencao Produtos\db"
embedding = HuggingFaceEmbeddings()
vector_store = pg_vector_connector._get_vector_store(
    collection_name=COLLECTION_NAME_PLANILHA,
)

# Função para remover miligramagem e outras unidades de medida dinâmicas da string do princípio ativo
def remover_miligramagem(principio_ativo):
    # Expressão regular para remover qualquer número seguido de unidades de medida (MG, g, ml, etc.)
    # Captura números com ou sem casas decimais seguidos de uma unidade de medida (ex: "MG", "g", "ml")
    return re.sub(r'\s?\d+(\.\d+)?\s*(mg|g|ml|mcg|kg|mcg|oz|ml|tablet|cap)', '', principio_ativo, flags=re.IGNORECASE).strip()

# Função para extrair os princípios ativos, levando em conta a possibilidade de múltiplos princípios
def extrair_principios_ativos(principio_ativo_completo):
    # Divida a string pelo símbolo "+"
    principios = principio_ativo_completo.split('+')
    
    # Remova a miligramagem de cada princípio ativo
    principios_limpos = [remover_miligramagem(p.strip()) for p in principios]
    
    return principios_limpos

def busca_principio_ativo(principios_ativos: List[str], current_vector_store: Any) -> List[Any]:
    """Realiza busca por similaridade e filtra por correspondência exata do princípio ativo."""
    all_retrieved_docs = set()
    for principio in principios_ativos:
        # 1. Realiza a busca e atribui os resultados a uma variável
        retrieved_docs = current_vector_store.similarity_search(principio, k=20)
        
        # 2. Filtra esta lista de documentos
        filtered_docs = [
            doc for doc in retrieved_docs 
            if principio.lower() in doc.page_content.lower()
        ]
        
        # 3. Adiciona os documentos filtrados ao conjunto para garantir unicidade
        for doc in filtered_docs:
            all_retrieved_docs.add(doc)
    
    # 4. Retorna a lista de documentos únicos
    return list(all_retrieved_docs)

# Função para buscar documentos relacionados ao NCM
def buscar_por_ncm(ncm_usuario, current_vector_store):
    # Realiza a busca de similaridade com o NCM fornecido
    results = current_vector_store.similarity_search(ncm_usuario, k=20)
    
    # Filtra os documentos que realmente possuem o NCM desejado
    resultados_filtrados = []
    for doc in results:
        ncm_relacionado = doc.metadata.get("ncm", "").split(",")
        if ncm_usuario in ncm_relacionado:
            resultados_filtrados.append(doc)
    
    return resultados_filtrados

system_prompt = '''
Você é um assistente especializado em isenção de ICMS.
Sua única e exclusiva tarefa é analisar o contexto fornecido para determinar se um produto é isento de ICMS e, se for, identificar o convênio ICMS vinculado.

**Contexto Fornecido:**
Cada entrada no contexto descreve uma isenção. O formato é: "Isenção: [Nome da Isenção]. Princípio Ativo: [Princípio Ativo]. Observação: [Observação]. NCMs relacionados: [Lista de NCMs]."
A parte "[Nome da Isenção]" contém o número do convênio (ex: Convênio 126/10).

**Instruções de Análise e Resposta (CRÍTICO - SIGA A ORDEM RIGOROSAMENTE):**
1.  **PRIMEIRO - Busque pelo Princípio Ativo:**
    * Examine a pergunta do usuário. Se o produto tiver um "Princípio Ativo" (ex: OLAPARIBE 150 MG), ignore a miligramagem (ou qualquer outra quantidade) e considere APENAS o nome do princípio ativo (ex: OLAPARIBE).
    * Procure por este nome do princípio ativo no campo "Princípio Ativo" dos documentos do contexto.
    * **Se encontrar uma correspondência válida pelo Princípio Ativo, pare de procurar e use esta informação.**
    * **Se houver mais de um princípio ativo no produto, AMBOS devem estar presentes no mesmo convênio para ser considerado isento.**
2.  **SEGUNDO - Busque pelo NCM (Somente se o Princípio Ativo NÃO foi encontrado):**
    * Se você NÃO encontrou uma correspondência pelo Princípio Ativo (ou se o Princípio Ativo não foi fornecido na pergunta), então, e somente então, verifique se o NCM do produto (ex: 30049069) está presente na lista de "NCMs relacionados" em qualquer documento do contexto.
    * **Se encontrar uma correspondência válida pelo NCM, use esta informação.**
    Condição de isenção pelo NCM:
    * **Caso o NCM relacionado tenha um princípio ativo associado, você deve considerar a combinação de princípio ativo + NCM para retornar se o produto que passamos é isento ou não** *
    * **Caso o NCM relacionado não tenha um princípio ativo associado, aí sim você pode considerar o produto como isento** *
3.  **Condição de Isenção:** Se, e somente se, uma correspondência foi encontrada por qualquer um dos métodos acima (Princípio Ativo ou NCM), o produto é considerado isento de ICMS.
4.  **Formato da Resposta (MUITO IMPORTANTE - Siga ESTE FORMATO EXATAMENTE):**
    * **Se o produto for isento:**
        * Responda: "Sim. Convênio ICMS [NÚMERO_DO_CONVÊNIO]".
        * O [NÚMERO_DO_CONVÊNIO] deve ser extraído da parte "Isenção: Convênio [NÚMERO_DO_CONVÊNIO]" do documento correspondente encontrado. Por exemplo, se o contexto diz "Isenção: Convênio 126/10", você deve usar "126/10".
        * Certifique-se de adicionar "ICMS" entre "Convênio" e o número.
    * **Se o produto NÃO for isento (se nenhuma correspondência válida foi encontrada pelo Princípio Ativo nem pelo NCM):**
        * Responda: "Não".

**Exemplos de Resposta Esperada:**
* "Sim. Convênio ICMS 126/10"
* "Sim. Convênio ICMS 10/02"
* "Sim. Convênio ICMS 162/94 - 132/2021"
* "Não"

Me retorne apenas a resposta direta, sem mais informações ou explicações adicionais.

Contexto: {context}
'''

prompt = ChatPromptTemplate.from_messages(
    [
        ("system", system_prompt),
        ("human", "{input}"),
    ]
)

# Cria um chain já pré preenchido (feito pela comunidade do langchain)
question_answer_chain = create_stuff_documents_chain(
    llm=model,
    prompt=prompt,
)

def obter_resposta_base_tool(input_str: str) -> str:
    """
    Obtém a resposta de isenção de ICMS para um produto.
    O input_str deve ser uma string JSON contendo 'codigo', 'nome', 'principio_ativo' e 'ncm'.
    Exemplo: '{"codigo": "123", "nome": "Produto X", "principio_ativo": "OLAPARIBE 150 MG", "ncm": "30049069"}'
    """
    try:
        product_data = json.loads(input_str)
        codigo = product_data.get("codigo")
        nome = product_data.get("nome")
        principio_ativo = product_data.get("principio_ativo")
        ncm = product_data.get("ncm")

        retrieved_docs = []
        
        if principio_ativo:
            principios_ativos = extrair_principios_ativos(principio_ativo)
            print(f"  [Tool] Buscando por princípios ativos: {principios_ativos}")
            retrieved_docs = busca_principio_ativo(principios_ativos, vector_store)
        
        if not retrieved_docs and ncm:
            print(f"  [Tool] Nenhum documento encontrado por princípio ativo ou não fornecido; buscando por NCM: {ncm}")
            retrieved_docs = buscar_por_ncm(ncm, vector_store)
        
        if not retrieved_docs:
            print("  [Tool] Nenhum documento relevante encontrado para a consulta.")
            return "Não"
        
        print(f"  [Tool] Encontrados {len(retrieved_docs)} documentos relevantes.")

        query = f"""
        Verifique o contexto se o produto é isento de ICMS e se sim, me retorne o convênio vinculado. Analise primeiro pelo NCM, caso não encontre, analise pelo princípio ativo (o princípio ativo será destacado no produto que vou listar pra você). Me retorne apenas 'Sim' ou 'Não' e o convênio, sem mais informações.
        Pode ser que o produto não tenha princípio ativo ou NCM, nesse caso, verifique apenas a informação que está presente no produto.
        Se houver mais de um convênio, retorne todos os convênios separados por vírgula.

        Produto:
        Código: {codigo}
        Princípio ativo: {principio_ativo}
        NCM: {ncm}
        Nome: {nome}
        """
        
        response = question_answer_chain.invoke(
            {"input": query, "context": retrieved_docs}
        )
        return response
    
    except json.JSONDecodeError:
        return "Erro: O input para a ferramenta obter_resposta_base_tool deve ser um JSON válido."
    except Exception as e:
        return f"Erro inesperado na ferramenta obter_resposta_base_tool: {e}"
    
    
# 1. Criar a ferramenta a partir da função
get_icms_exemption_tool = tools.Tool(
    name="get_icms_exemption",
    func=obter_resposta_base_tool,
    description="Útil para determinar se um produto é isento de ICMS e qual convênio está vinculado. O input deve ser uma string JSON com 'codigo', 'nome', 'principio_ativo', 'ncm'."
)

# Lista de ferramentas que o agente pode usar
tools_for_agent = [get_icms_exemption_tool]

# 2. Definir o Prompt para o Agente
agent_template_string = """
Você é um assistente que determina a isenção de ICMS de produtos e formata a resposta em JSON.

**Sua TAREFA ÚNICA E FINAL é:**
1. Usar a ferramenta 'get_icms_exemption' EXATAMENTE uma vez.
2. Pegar a resposta da ferramenta (ex: 'Sim. Convênio ICMS 126/10' ou 'Não').
3. Converter ESSA resposta para o formato JSON final e fornecer APENAS ESSE JSON como sua FINAL ANSWER.

**ATENÇÃO: A entrada para a ferramenta 'get_icms_exemption' deve ser uma string JSON válida, sem aspas extras.**

Available tools:
{tools}
Ferramentas que você pode usar: {tool_names} 

Use o seguinte formato de raciocínio ReAct:

Question: A pergunta ou tarefa de entrada.
Thought: Você deve sempre pensar no que fazer. Minha primeira ação DEVE ser chamar a ferramenta 'get_icms_exemption' com o input JSON.
Action: A ação a ser tomada, deve ser uma de [{tool_names}].
Action Input: # O JSON deve ser colocado sem aspas extras. Por exemplo: {{"codigo": "123", ...}}
Observation: O resultado da ação.
Thought: Eu obtive a resposta da ferramenta. Minha ÚNICA tarefa agora é formatar a resposta no JSON e dar a Final Answer.
Final Answer: # O objeto JSON virá aqui. NADA MAIS.

Começar!

Question: {input}
Thought: {agent_scratchpad}
"""
agent_prompt = PromptTemplate.from_template(agent_template_string)

# 3. Criar o Agente
agent = create_react_agent(llm=model, tools=tools_for_agent, prompt=agent_prompt) # Usar llm_model aqui
agent_executor = AgentExecutor(agent=agent, tools=tools_for_agent, verbose=True, handle_parsing_errors=True, max_iterations=15)

# --- CÓDIGO PRINCIPAL PARA RODAR A AUTOMAÇÃO ---
if __name__ == "__main__":
    caminho_arquivo = r"C:\Users\Nícolas Nasário\Downloads\Base Teste.xlsx"
    workbook = load_workbook(caminho_arquivo)
    sheet = workbook["1"]

    for row_idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        current_sheet_row = row_idx + 2

        codigo = row_data[0]
        principio_ativo = row_data[1]
        ncm = row_data[2]
        nome = row_data[3]

        print(f"\n--- Processando produto (Linha {current_sheet_row}): Código={codigo}, PA={principio_ativo}, NCM={ncm} ---")
        
        # O input para o agente será uma descrição do produto
        # O agente então decidirá usar a ferramenta 'get_icms_exemption' com os dados formatados em JSON
        agent_input = {
            "codigo": codigo,
            "nome": nome,
            "principio_ativo": principio_ativo,
            "ncm": ncm
        }
        
        # O agente espera uma string como input, então convertemos o dicionário para JSON
        input_for_agent = json.dumps(agent_input)

        try:
            # Invocar o agente para obter a resposta formatada
            # O agente vai chamar a ferramenta 'get_icms_exemption' internamente
            agent_response = agent_executor.invoke({"input": input_for_agent})
            
            # A resposta do agente é um dicionário, que contém a chave 'output'
            resposta_formatada = agent_response.get('output', 'Erro: Resposta do agente não encontrada.')
            
            retorno_json_formatado = json.loads(resposta_formatada) 
            
            print(f"Resposta formatada pelo agente para o produto na linha {current_sheet_row}: {resposta_formatada}")

            if "não" in str(retorno_json_formatado.get("status_isencao")).lower():
                # Salva a resposta na coluna 5 da planilha
                sheet.cell(row=current_sheet_row, column=5).value = f"""{retorno_json_formatado.get("status_isencao")}."""
            else:
                # Salva a resposta na coluna 5 da planilha
                sheet.cell(row=current_sheet_row, column=5).value = f"""{retorno_json_formatado.get("status_isencao")}. Convênio {retorno_json_formatado.get("convenio_icms", "")}"""
            
        except Exception as e:
            print(f"Erro ao processar produto com o agente na linha {current_sheet_row}: {e}")
            sheet.cell(row=current_sheet_row, column=5).value = f"Erro: {e}"
            
    workbook.save(caminho_arquivo)
    workbook.close()
    print("\nProcessamento da planilha 'Base Teste.xlsx' concluído e resultados salvos.")


# caminho_arquivo = r"C:\Users\Nícolas Nasário\Downloads\Base Teste.xlsx"
# workbook = load_workbook(caminho_arquivo)
# sheet = workbook["1"]

# linha_inicial = 2
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     codigo = row[0]
#     principio_ativo = row[1]
#     ncm = row[2]
#     nome = row[3]

#     # Obter a resposta
#     resposta = obter_resposta_com_base(codigo, nome, principio_ativo, ncm, vector_store)
#     print(resposta)

#     sheet.cell(row=linha_inicial, column=5).value = resposta
#     linha_inicial += 1

    
# workbook.save(caminho_arquivo)
# workbook.close()