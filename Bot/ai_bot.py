import os
import re
from openpyxl import Workbook, load_workbook
from langchain_groq import ChatGroq
from langchain_chroma import Chroma
from langchain.chains.combine_documents import create_stuff_documents_chain # Retorna um chain pronto
from langchain.chains.retrieval import create_retrieval_chain # Chain que usa o retriever (busca nos dados)
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_core.prompts import ChatPromptTemplate
from dotenv import load_dotenv


load_dotenv()
groq_api_key = os.getenv("GROQ_API_KEY")

model = ChatGroq(
    model = "meta-llama/llama-4-maverick-17b-128e-instruct",
)

# Cria uma conexão com o banco de dados de vetores
persist_directory = rf"C:\rpa\Python\Modelo Isencao Produtos\db"
embedding = HuggingFaceEmbeddings()
vector_store = Chroma(
    persist_directory=persist_directory,
    embedding_function=embedding,
    collection_name="planilha_data",
)

# Função para remover miligramagem e outras unidades de medida dinâmicas da string do princípio ativo
def remover_miligramagem(principio_ativo):
    # Expressão regular para remover qualquer número seguido de unidades de medida (MG, g, ml, etc.)
    # Captura números com ou sem casas decimais seguidos de uma unidade de medida (ex: "MG", "g", "ml")
    return re.sub(r'\s?\d+(\.\d+)?\s*(mg|g|ml|mcg|kg|mcg|oz|ml|tablet|cap)', '', principio_ativo, flags=re.IGNORECASE).strip()

# Função para extrair os princípios ativos, levando em conta a possibilidade de múltiplos princípios
def extrair_principios_ativos(principio_ativo_completo):
    # Divida a string pelo símbolo "+"
    principios = principio_ativo_completo.split('+')
    
    # Remova a miligramagem de cada princípio ativo
    principios_limpos = [remover_miligramagem(p.strip()) for p in principios]
    
    return principios_limpos

# Criação de um método para buscar dinamicamente o princípio ativo
def busca_principio_ativo(principios_ativos):
    # Para cada princípio ativo, realiza a busca de similaridade
    results = []
    for principio in principios_ativos:
        results += vector_store.similarity_search(principio, k=20)
    return results

# Função para buscar documentos relacionados ao NCM
def buscar_por_ncm(ncm_usuario):
    # Realiza a busca de similaridade com o NCM fornecido
    results = vector_store.similarity_search(ncm_usuario, k=20)
    
    # Filtra os documentos que realmente possuem o NCM desejado
    resultados_filtrados = []
    for doc in results:
        ncm_relacionado = doc.metadata.get("ncm", "").split(",")
        if ncm_usuario in ncm_relacionado:
            resultados_filtrados.append(doc)
    
    return resultados_filtrados

system_prompt = '''
Você é um assistente especializado em isenção de ICMS.
Sua única e exclusiva tarefa é analisar o contexto fornecido para determinar se um produto é isento de ICMS e, se for, identificar o convênio ICMS vinculado.

**Contexto Fornecido:**
Cada entrada no contexto descreve uma isenção. O formato é: "Isenção: [Nome da Isenção]. Princípio Ativo: [Princípio Ativo]. Observação: [Observação]. NCMs relacionados: [Lista de NCMs]."
A parte "[Nome da Isenção]" contém o número do convênio (ex: Convênio 126/10).

**Instruções de Análise e Resposta (CRÍTICO - SIGA A ORDEM RIGOROSAMENTE):**
1.  **PRIMEIRO - Busque pelo Princípio Ativo:**
    * Examine a pergunta do usuário. Se o produto tiver um "Princípio Ativo" (ex: OLAPARIBE 150 MG), ignore a miligramagem (ou qualquer outra quantidade) e considere APENAS o nome do princípio ativo (ex: OLAPARIBE).
    * Procure por este nome do princípio ativo no campo "Princípio Ativo" dos documentos do contexto.
    * **Se encontrar uma correspondência válida pelo Princípio Ativo, pare de procurar e use esta informação.**
2.  **SEGUNDO - Busque pelo NCM (Somente se o Princípio Ativo NÃO foi encontrado):**
    * Se você NÃO encontrou uma correspondência pelo Princípio Ativo (ou se o Princípio Ativo não foi fornecido na pergunta), então, e somente então, verifique se o NCM do produto (ex: 30049069) está presente na lista de "NCMs relacionados" em qualquer documento do contexto.
    * **Se encontrar uma correspondência válida pelo NCM, use esta informação.**
3.  **Condição de Isenção:** Se, e somente se, uma correspondência foi encontrada por qualquer um dos métodos acima (Princípio Ativo ou NCM), o produto é considerado isento de ICMS.
4.  **Formato da Resposta (MUITO IMPORTANTE - Siga ESTE FORMATO EXATAMENTE):**
    * **Se o produto for isento:**
        * Responda: "Sim. Convênio ICMS [NÚMERO_DO_CONVÊNIO]".
        * O [NÚMERO_DO_CONVÊNIO] deve ser extraído da parte "Isenção: Convênio [NÚMERO_DO_CONVÊNIO]" do documento correspondente encontrado. Por exemplo, se o contexto diz "Isenção: Convênio 126/10", você deve usar "126/10".
        * Certifique-se de adicionar "ICMS" entre "Convênio" e o número.
    * **Se o produto NÃO for isento (se nenhuma correspondência válida foi encontrada pelo Princípio Ativo nem pelo NCM):**
        * Responda: "Não".

**Exemplos de Resposta Esperada:**
* "Sim. Convênio ICMS 126/10"
* "Sim. Convênio ICMS 10/02"
* "Sim. Convênio ICMS 162/94 - 132/2021"
* "Não"

Me retorne apenas a resposta direta, sem mais informações ou explicações adicionais.

Contexto: {context}
'''

prompt = ChatPromptTemplate.from_messages(
    [
        ("system", system_prompt),
        ("human", "{input}"),
    ]
)

# Cria um chain já pré preenchido (feito pela comunidade do langchain)
question_answer_chain = create_stuff_documents_chain(
    llm=model,
    prompt=prompt,
)

# Função para consultar e obter a resposta com base no princípio ativo ou NCM
def obter_resposta_com_base(codigo, nome, principio_ativo, ncm):
    # Busca os resultados com base no princípio ativo fornecido
    if principio_ativo:
        # Extrair os princípios ativos
        principios_ativos = extrair_principios_ativos(principio_ativo)
        results = busca_principio_ativo(principios_ativos)
    else:
        # Se o princípio ativo não for fornecido, busca pelo NCM
        results = buscar_por_ncm(ncm)
    
    if len(results) == 0:
        return "Não"
    
    else:
        # Prepara a consulta
        query = f"""
        Verifique o contexto se o produto é isento de ICMS e se sim, me retorne o convênio vinculado. Analise primeiro pelo NCM, caso não encontre, analise pelo princípio ativo (o princípio ativo será destacado no produto que vou listar pra você). Me retorne apenas 'Sim' ou 'Não' e o convênio, sem mais informações.
        Pode ser que o produto não tenha princípio ativo ou NCM, nesse caso, verifique apenas a informação que está presente no produto.
        Se houver mais de um convênio, retorne todos os convênios separados por vírgula.

        Produto:
        Código: {codigo}
        Princípio ativo: {principio_ativo}
        NCM: {ncm}
        Nome: {nome}
        """
        
        # Invoca a chain para obter a resposta
        response = question_answer_chain.invoke(
            {"input": query, "context": results}
        )

        return response


caminho_arquivo = r"C:\Users\Nícolas Nasário\Downloads\Base Teste.xlsx"
workbook = load_workbook(caminho_arquivo)
sheet = workbook["1"]

linha_inicial = 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    codigo = row[0]
    principio_ativo = row[1]
    ncm = row[2]
    nome = row[3]

    # Obter a resposta
    resposta = obter_resposta_com_base(codigo, nome, principio_ativo, ncm)
    print(resposta)

    workbook = load_workbook(caminho_arquivo)
    sheet.cell(row=linha_inicial, column=5).value = resposta
    linha_inicial += 1

    
    workbook.save(caminho_arquivo)
    workbook.close()