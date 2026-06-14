# rag_core.py
import os
import re
import sys
import json
import unicodedata
from datetime import datetime
from dotenv import load_dotenv, find_dotenv
from openpyxl import load_workbook
from typing import Optional, Any, Dict, List

# Langchain imports
from langchain_groq import ChatGroq
from langchain_openai import ChatOpenAI
from langchain.chains.combine_documents import create_stuff_documents_chain
from langchain_core.prompts import ChatPromptTemplate, PromptTemplate
from langchain.agents import create_react_agent, AgentExecutor
from langchain import tools

# Ensure your custom ConectaPGVector class is in PYTHONPATH or properly imported
# Ajuste este caminho se ConectaPGVector.py estiver em outro lugar
sys.path.append(r"C:\rpa\Python") 
from Classes.Postgres.Postgres.ConectaPGVector import ConectaPGVector
from Classes.MoverArquivos.MoverArquivos.HubArquivos import HubArquivos
from Classes.Gmail.Gmail.ConectaGmail import ConectaGmail
from Classes.ZimbraMailer.ZimbraMailer.Zimbra import ZimbraMailer

# Obtém o caminho do diretório onde o script está localizado
script_dir = os.path.dirname(os.path.abspath(__file__))
# Procura o .env a partir do diretório do script
dotenv_path = find_dotenv(os.path.join(script_dir, '.env'))
# Carrega o .env
load_dotenv(dotenv_path)

groq_api_key = os.getenv("GROQ_API_KEY")
openai_api_key = os.getenv("OPENAI_API_KEY")

# Adicionado: Token Secreto para a API
API_SECRET_TOKEN = os.getenv("API_SECRET_TOKEN") 
if API_SECRET_TOKEN is None:
    raise ValueError("API_SECRET_TOKEN não configurado nas variáveis de ambiente. Por favor, adicione-o ao seu arquivo .env.")

COLLECTION_NAME_PLANILHA = "isencoes_de_produtos"

# --- Componentes de IA (Serão Inicializados uma única vez na função initialize_rag_components) ---
# Usamos Optional para indicar que podem ser None antes da inicialização
# llm_model: Optional[ChatGroq] = None
llm_model: Optional[ChatOpenAI] = None
raw_pg_vector_store: Optional[Any] = None # Tipo Any para PGVector
agent_executor: Optional[AgentExecutor] = None

# --- Funções Auxiliares de Pré-processamento ---
def remover_miligramagem(principio_ativo: str) -> str:
    """Remove unidades de miligramagem e outras da string do princípio ativo."""
    return re.sub(r'\s?\d+(\.\d+)?\s*(mg/ml|mg|g|ml|mcg|kg|mcg|oz|ml|tablet|cap)', '', principio_ativo, flags=re.IGNORECASE).strip()


# --- Função para remover o ânion ---
def remover_anion(principio_ativo: str) -> str:
    """Remove unidades de miligramagem e outras da string do princípio ativo."""
    return re.sub(r'\s*(cloreto|sulfato|nitrato|fosfato|bicarbonato|acetato|tiossulfato|lactato|citrato|hidróxido|carbonato|borato|fluoreto|iodeto|brometo|peróxido|sulfeto|manganato|permanganato|cianeto|tartato|fosfonato|tetraborato|ácido sulfidrico|hidrogênico|oxalato|fenilalanato|ácido metanosulfônico|ácido acético|acetato de cálcio|etilenodiaminotetraacetato|dicromato|ácido clorosulfônico|bifluoreto|sulfito|metabisulfito|ácido fosfônico|sulfato de alumínio|ácido tartárico|ácido maleico|ácido fumárico|cloridrato|dihidratado|dihidratada|di-hidratado|di-hidratada|trihidratado|trihidratada|tri-hidratado|tri-hidratada|hidratado|hidratada|de)', '', principio_ativo, flags=re.IGNORECASE).strip()


def extrair_principios_ativos(principio_ativo_completo: str) -> List[str]:
    """Extrai e limpa múltiplos princípios ativos de uma string, considerando diferentes delimitadores."""
    if not principio_ativo_completo:
        return []
    
    # Primeiro, normaliza a string removendo espaços extras
    principio_ativo_completo = re.sub(r'\s+', ' ', principio_ativo_completo.strip())
    
    # Lista de delimitadores possíveis baseados na análise da planilha
    delimitadores = [
        r'\s*\+\s*',      # Sinal de +
        r'\s*;\s*',       # Ponto e vírgula
        r'\s*,\s*e\s*',   # Vírgula seguida de "e"
        r'\s*\bou\s*',    # Palavra "ou"
    ]
    
    # Tenta dividir usando cada delimitador
    principios = [principio_ativo_completo]
    for delimitador in delimitadores:
        novos_principios = []
        for principio in principios:
            if re.search(delimitador, principio):
                partes = re.split(delimitador, principio)
                novos_principios.extend(partes)
            else:
                novos_principios.append(principio)
        principios = novos_principios
    
    # Limpa cada princípio ativo
    principios_limpos = []
    for p in principios:
        p_limpo = p.strip()
        if p_limpo:
            # p_limpo = remover_anion(remover_miligramagem(p_limpo))
            p_limpo = remover_miligramagem(p_limpo)
            if p_limpo:  # Só adiciona se ainda há conteúdo após limpeza
                principios_limpos.append(p_limpo)
    
    return principios_limpos


def busca_principio_ativo(principios_ativos: List[str], current_vector_store: Any) -> List[Any]:
    """Realiza busca por similaridade considerando diferentes estratégias para princípios ativos compostos."""
    all_retrieved_docs = {}
    
    # Estratégia 1: Busca pela combinação completa original
    principio_completo = " + ".join(principios_ativos)
    retriever = current_vector_store.as_retriever(search_type="mmr", search_kwargs={"k": 4, "fetch_k": 20})
    retrieved_docs = retriever.invoke(principio_completo)
    
    for doc in retrieved_docs:
        doc_key = (doc.page_content, tuple(doc.metadata.items()))
        all_retrieved_docs[doc_key] = doc
    
    # Estratégia 2: Busca por cada princípio ativo individualmente
    for principio in principios_ativos:
        retrieved_docs = retriever.invoke(principio)
        for doc in retrieved_docs:
            doc_key = (doc.page_content, tuple(doc.metadata.items()))
            all_retrieved_docs[doc_key] = doc
    
    # Estratégia 3: Busca por combinações invertidas (ex: "B + A" quando temos "A + B")
    if len(principios_ativos) > 1:
        principios_invertidos = principios_ativos[::-1]
        principio_invertido = " + ".join(principios_invertidos)
        retrieved_docs = retriever.invoke(principio_invertido)
        for doc in retrieved_docs:
            doc_key = (doc.page_content, tuple(doc.metadata.items()))
            all_retrieved_docs[doc_key] = doc
    
    return list(all_retrieved_docs.values())


def buscar_por_ncm(ncm_usuario: str, current_vector_store: Any) -> List[Any]:
    """Realiza busca por similaridade de NCM e filtra resultados."""
    # results = current_vector_store.similarity_search(ncm_usuario, k=20)

    # 1. Obtém o retriever
    retriever = current_vector_store.as_retriever(search_type="mmr", search_kwargs={"k": 4, "fetch_k": 20})
    # 2. Invoca o retriever para obter os documentos
    results = retriever.invoke(ncm_usuario)

    resultados_filtrados = []
    for doc in results:
        ncm_metadata_str = doc.metadata.get("ncm", "")
        ncm_list_in_doc = [n.strip() for n in ncm_metadata_str.split(',') if n.strip()]
        if ncm_usuario in ncm_list_in_doc:
            resultados_filtrados.append(doc)
    return resultados_filtrados


def normalize(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII").casefold()


# def checar_correspondencia_completa(principios_produto: List[str], docs: List[Any]) -> Optional[Any]:
#     """Verifica correspondência completa considerando diferentes formatos de princípios ativos compostos."""
#     principios_produto_norm = [normalize(p) for p in principios_produto]
    
#     for doc in docs:
#         principio_doc = doc.metadata.get("principio_ativo", "").strip()
        
#         if not principio_doc:
#             continue
            
#         # Extrai princípios ativos do documento usando a função aprimorada
#         principios_doc_lista = extrair_principios_ativos(principio_doc)
#         principios_doc_norm = [normalize(p) for p in principios_doc_lista]
        
#         # Verifica correspondência exata (todos os princípios presentes)
#         if len(principios_produto_norm) == len(principios_doc_norm):
#             # Verifica se todos os princípios do produto estão no documento (mesma ordem)
#             if all(p_prod in principios_doc_norm for p_prod in principios_produto_norm):
#                 return doc
            
#             # Verifica se todos os princípios estão presentes (ordem diferente)
#             if set(principios_produto_norm) == set(principios_doc_norm):
#                 return doc
        
#         # Verifica se todos os princípios do produto estão contidos no texto completo do documento
#         principio_doc_completo_norm = normalize(principio_doc)
#         if all(p_prod in principio_doc_completo_norm for p_prod in principios_produto_norm):
#             return doc
    
#     return None


def checar_correspondencia_completa(principios_produto: List[str], docs: List[Any]) -> Optional[Any]:
    """Versão corrigida da função atual."""
    principios_produto_norm = [normalize(p) for p in principios_produto]
    
    for doc in docs:
        principio_doc = doc.metadata.get("principio_ativo", "").strip()
        if not principio_doc:
            continue
            
        principios_doc_lista = extrair_principios_ativos(principio_doc)
        principios_doc_norm = [normalize(p) for p in principios_doc_lista]
        
        # Verifica correspondência exata (mesmo número de princípios)
        if len(principios_produto_norm) == len(principios_doc_norm):
            if set(principios_produto_norm) == set(principios_doc_norm):
                return doc
        
        # NOVA VERIFICAÇÃO: Correspondência bidirecional com substring
        # Verifica se cada princípio do produto tem correspondência no documento
        produto_match = all(
            any(p_prod in p_doc or p_doc in p_prod for p_doc in principios_doc_norm) 
            for p_prod in principios_produto_norm
        )
        
        # Verifica se cada princípio do documento tem correspondência no produto
        doc_match = all(
            any(p_doc in p_prod or p_prod in p_doc for p_prod in principios_produto_norm) 
            for p_doc in principios_doc_norm
        )
        
        if produto_match and doc_match:
            return doc
    
    return None


def verificar_principios_mesmo_convenio(principios_ativos: List[str], docs: List[Any]) -> Optional[str]:
    """
    Verifica se todos os princípios ativos estão presentes no mesmo convênio,
    mesmo que em documentos/registros separados.
    """
    if len(principios_ativos) <= 1:
        return None
    
    # Agrupa documentos por convênio
    convenios = {}
    for doc in docs:
        convenio = doc.metadata.get("nome_isencao", "")
        if convenio:
            if convenio not in convenios:
                convenios[convenio] = []
            convenios[convenio].append(doc)
    
    # Para cada convênio, verifica se todos os princípios estão presentes
    principios_normalizados = [normalize(p) for p in principios_ativos]
    
    for convenio, docs_convenio in convenios.items():
        principios_encontrados = set()
        
        for doc in docs_convenio:
            principio_doc = doc.metadata.get("principio_ativo", "").strip()
            if principio_doc:
                principio_doc_norm = normalize(principio_doc)
                
                # Verifica quais princípios do produto estão neste documento
                for principio in principios_normalizados:
                    if principio_doc_norm in principio:
                        principios_encontrados.add(principio)
        
        # Se todos os princípios foram encontrados neste convênio
        if len(principios_encontrados) == len(principios_normalizados):
            return convenio
    
    return None


# --- Ferramenta e Lógica do Agente ---
def _obter_resposta_qa_chain(query: str, context: List[Any], llm_model_instance: ChatOpenAI) -> str:
    """Função interna para invocar a cadeia de QA."""
    qa_system_prompt = """
    Você é um assistente especializado em isenção de ICMS.
    Sua única e exclusiva tarefa é analisar o contexto fornecido para determinar se um produto é isento de ICMS e, se for, identificar o convênio ICMS vinculado.

    **Contexto Fornecido:**
    Cada entrada no contexto descreve uma isenção. O formato é:
    "Isenção: [Nome da Isenção]. Princípio Ativo: [Princípio Ativo]. Observação: [Observação]. NCMs relacionados: [Lista de NCMs]."
    A parte "[Nome da Isenção]" contém o número do convênio (ex: Convênio 126/10).

    **Instruções de Análise e Resposta (SIGA RIGOROSAMENTE):**

    1. **PRIMEIRA REGRA – Análise por Princípio Ativo:**
    * Para produtos com um único princípio ativo: verifique se aparece no campo "Princípio Ativo" do documento.
    * Para produtos com múltiplos princípios ativos (ex: A + B):
        - Primeiro, procure pela combinação exata (A + B) no documento;
        - Se não encontrar, procure pela combinação invertida (B + A);
        - Se não encontrar, verifique se todos os princípios individuais (A e B) estão presentes no mesmo documento;
        - Se não encontrar, verifique se há documentos separados para cada princípio, mas que façam parte do mesmo convênio.
    * Se encontrar correspondência válida, o produto é isento. Use a informação deste documento para a resposta final.

    2. **SEGUNDA REGRA – Análise por NCM (condição de vazio):**
    * Só utilize esta regra se o princípio ativo do produto for VAZIO **e** o princípio ativo do documento também for VAZIO.
    * Nesse caso, compare apenas o NCM do produto com os NCMs listados no documento.
    * Se houver correspondência, o produto é isento.

    3. **Condição Final de Isenção:**
    * Se encontrou uma isenção válida em qualquer uma das duas regras acima, o produto é isento.
    * Caso contrário, o produto NÃO é isento.

    4. **Formato da Resposta (OBRIGATÓRIO):**
    * Se o produto for isento:
        - Responda: "Sim. Convênio ICMS [NÚMERO_DO_CONVÊNIO]"
    * Se o produto NÃO for isento:
        - Responda: "Não"

    **Exemplos de Resposta Esperada:**
    * "Sim. Convênio ICMS 126/10"
    * "Sim. Convênio ICMS 10/02"
    * "Sim. Convênio ICMS 162/94 - 132/2021"
    * "Não"

    Contexto: {context}
    """
    qa_prompt = ChatPromptTemplate.from_messages([("system", qa_system_prompt), ("human", "{input}"),])
    qa_chain = create_stuff_documents_chain(llm=llm_model_instance, prompt=qa_prompt,)
    return qa_chain.invoke({"input": query, "context": context})


def _obter_resposta_base_tool_func(input_str: str) -> str:
    """
    Função que atua como ferramenta do Agente.
    Converte dados JSON de entrada, busca no vector store e invoca a QA Chain.
    Utiliza as variáveis globais `raw_pg_vector_store` e `llm_model`.
    """
    if raw_pg_vector_store is None or llm_model is None:
        raise RuntimeError("Componentes RAG não inicializados. Chame initialize_rag_components primeiro.")

    try:
        product_data = json.loads(input_str)
        codigo = product_data.get("codigo")
        nome = product_data.get("nome")
        principio_ativo = product_data.get("principio_ativo")
        ncm = product_data.get("ncm")

        retrieved_docs = []
        
        # if principio_ativo:
        #     principios_ativos = extrair_principios_ativos(principio_ativo)
        #     retrieved_docs = busca_principio_ativo(principios_ativos, raw_pg_vector_store)
            
        #     # Verifica correspondência completa primeiro
        #     matched_doc = checar_correspondencia_completa(principios_ativos, retrieved_docs)
        #     if matched_doc:
        #         conv = matched_doc.metadata.get("nome_isencao")
        #         return f"Sim. Convênio ICMS {conv}"
            
        #     # NOVA VERIFICAÇÃO: Se há múltiplos princípios, verifica se estão no mesmo convênio
        #     if len(principios_ativos) > 1:
        #         convenio_encontrado = verificar_principios_mesmo_convenio(principios_ativos, retrieved_docs)
        #         if convenio_encontrado:
        #             return f"Sim. Convênio ICMS {convenio_encontrado}"
            
        #     # Se não encontrou correspondência exata nem no mesmo convênio, usa o LLM
        #     else:
        #         # Aqui entra o LLM para analisar múltiplos princípios ativos
        #         if retrieved_docs:
        #             query = f"Verifique se o produto com princípios ativos {principios_ativos} é isento de ICMS considerando os documentos fornecidos."
        #             return _obter_resposta_qa_chain(query, retrieved_docs, llm_model)
        #         else:
        #             if ncm:
        #                 retrieved_docs = buscar_por_ncm(ncm, raw_pg_vector_store)

        if principio_ativo:
            principios_ativos = extrair_principios_ativos(principio_ativo)
            retrieved_docs = busca_principio_ativo(principios_ativos, raw_pg_vector_store)
            
            # Verifica correspondência completa primeiro
            matched_doc = checar_correspondencia_completa(principios_ativos, retrieved_docs)
            if matched_doc:
                conv = matched_doc.metadata.get("nome_isencao")
                return f"Sim. Convênio ICMS {conv}"
            
            # NOVA VERIFICAÇÃO: Se há múltiplos princípios, verifica se estão no mesmo convênio
            if len(principios_ativos) > 1:
                convenio_encontrado = verificar_principios_mesmo_convenio(principios_ativos, retrieved_docs)
                if convenio_encontrado:
                    return f"Sim. Convênio ICMS {convenio_encontrado}"
            
            # Se não encontrou correspondência exata nem no mesmo convênio, usa o LLM
            if retrieved_docs:
                query = f"Verifique se o produto com princípios ativos {principios_ativos} é isento de ICMS considerando os documentos fornecidos."
                return _obter_resposta_qa_chain(query, retrieved_docs, llm_model)
        
        # print(retrieved_docs)
        
        query = f"""
        Verifique se o produto é isento de ICMS.

        IMPORTANTE: Para produtos com múltiplos princípios ativos:
        1. Procure primeiro pela combinação exata no documento
        2. Se não encontrar, procure pela combinação em ordem diferente
        3. Se não encontrar, verifique se todos os princípios individuais estão no mesmo documento/convênio
        4. Considere variações de nomenclatura e sinônimos

        Produto para analisar:
        - Princípios ativos: {principios_ativos}
        """
        
        response = _obter_resposta_qa_chain(query, retrieved_docs, llm_model)
        return response
    
    except json.JSONDecodeError:
        return "Erro: O input para a ferramenta obter_resposta_base_tool deve ser um JSON válido."
    except Exception as e:
        return f"Erro inesperado na ferramenta obter_resposta_base_tool: {e}"
    

def initialize_rag_components():
    """
    Inicializa todos os componentes pesados de RAG (LLM, PGVector, Agente).
    Esta função deve ser chamada apenas uma vez na inicialização da aplicação.
    """
    global llm_model, raw_pg_vector_store, agent_executor # Para modificar as variáveis globais

    print("Inicializando ConectaPGVector...")
    pg_vector_connector = ConectaPGVector()

    print(f"Obtendo instância direta do PGVector para a coleção: '{COLLECTION_NAME_PLANILHA}'...")
    raw_pg_vector_store = pg_vector_connector._get_vector_store(COLLECTION_NAME_PLANILHA)
    if raw_pg_vector_store is None:
        raise RuntimeError(f"Não foi possível obter a instância do PGVector para a coleção '{COLLECTION_NAME_PLANILHA}'. Verifique a conexão e o nome da coleção.")
    print("Instância do PGVector obtida com sucesso.")

    print("Inicializando LLM (Groq)...")
    # llm_model = ChatGroq(
    #     model="meta-llama/llama-4-maverick-17b-128e-instruct",
    # )
    llm_model = ChatOpenAI(
        # model="gpt-4.1-mini",
        # model="gpt-5-mini",
        model="gpt-4o-mini",
    )
    print("LLM inicializado.")

    # Criar a ferramenta a partir da função
    get_icms_exemption_tool = tools.Tool(
        name="get_icms_exemption",
        func=_obter_resposta_base_tool_func,
        description="Útil para determinar se um produto é isento de ICMS e qual convênio está vinculado. O input deve ser uma string JSON com 'codigo', 'nome', 'principio_ativo', 'ncm'."
    )
    tools_for_agent = [get_icms_exemption_tool]

    agent_template_string = """
    Você é um assistente que determina a isenção de ICMS de produtos e formata a resposta em JSON.

    **Sua TAREFA ÚNICA E FINAL é:**
    1. Usar a ferramenta 'get_icms_exemption' EXATAMENTE uma vez com o input da pergunta.
    2. Pegar a resposta da ferramenta (ex: 'Sim. 162/94 - 132/2021' ou 'Não').
    3. Converter ESSA resposta para o formato JSON final e fornecer APENAS ESSE JSON como sua FINAL ANSWER.

    **ATENÇÃO: A entrada para a ferramenta 'get_icms_exemption' deve ser uma string JSON válida, sem aspas extras.**
    **ATENÇÃO: A sua Final Answer DEVE ter apenas as chaves 'isento' e 'convenio'. Não use outras chaves.**
    **ATENÇÃO: Pense passo-a-passo e com calma.**

    Available tools:
    {tools}
    Ferramentas que você pode usar: {tool_names} 

    Use o seguinte formato de raciocínio ReAct:

    Question: A pergunta ou tarefa de entrada.
    Thought: Você deve sempre pensar no que fazer. Minha primeira ação DEVE ser chamar a ferramenta 'get_icms_exemption' com o input JSON.
    Action: A ação a ser tomada, deve ser uma de [{tool_names}].
    Action Input: # O JSON deve ser colocado sem aspas extras. Por exemplo: {{"codigo": "123", ...}}
    Observation: O resultado da ação.
    Thought: Eu obtive a resposta da ferramenta. Minha ÚNICA tarefa agora é formatar a resposta no JSON e dar a Final Answer.
    Final Answer: # O objeto JSON virá aqui. NADA MAIS.

    Começar!

    Question: {input}
    Thought: {agent_scratchpad}
    """
    agent_prompt = PromptTemplate.from_template(agent_template_string)
    # agent_prompt = agent_prompt.partial(tool_names=", ".join([tool.name for tool in tools_for_agent]))

    agent = create_react_agent(llm=llm_model, tools=tools_for_agent, prompt=agent_prompt)
    agent_executor = AgentExecutor(
        agent=agent, tools=tools_for_agent, verbose=True, # Definir como False para produção
        handle_parsing_errors=True, max_iterations=15,
    )
    
    print("Componentes RAG inicializados com sucesso!")
    return agent_executor # Retorna o executor do agente já inicializado


def process_product_with_rag(
    codigo: Optional[str],
    nome: Optional[str],
    principio_ativo: Optional[str],
    ncm: Optional[str]
    ) -> Dict[str, Any]:
    """
    Processa os dados de um produto usando o Agente de RAG e retorna o resultado formatado.
    Esta função espera que o `agent_executor` já tenha sido inicializado globalmente.
    """
    global agent_executor # Garante que estamos usando a variável global
    if agent_executor is None:
        # Se a inicialização não foi feita (ex: em um script que não é a API, chamar initialize_rag_components())
        # Ou, para a API, isto é um erro crítico.
        raise RuntimeError("Agent Executor não inicializado. Chame initialize_rag_components() primeiro.")

    agent_input_data = {
        # "codigo": codigo,
        # "nome": nome,
        "principio_ativo": principio_ativo,
        "ncm": ncm
    }
    input_for_agent = json.dumps(agent_input_data)

    try:
        agent_response = agent_executor.invoke({"input": input_for_agent})
        raw_output_string = agent_response.get('output', '{"error": "Agent output missing"}')
        
        parsed_response = {}
        try:
            parsed_response = json.loads(raw_output_string)
        except json.JSONDecodeError:
            json_match = re.search(r'\{.*?\}', raw_output_string, re.DOTALL)
            if json_match:
                json_str_candidate = json_match.group(0)
                try:
                    parsed_response = json.loads(json_str_candidate)
                except json.JSONDecodeError:
                    parsed_response = {"isento_icms": "Erro de Formato", "convenio": None}
            else:
                parsed_response = {"isento_icms": "Erro de Formato", "convenio": None}
        
        isento_status = parsed_response.get("isento_icms") or parsed_response.get("isento")
        
        if isento_status and isento_status.lower() == "não":
            final_response = {"isento_icms": "Não", "convenio": None}
        elif isento_status:
            final_response = {"isento_icms": "Sim", "convenio": parsed_response.get("convenio", parsed_response.get("convenio_icms"))}
        elif not isento_status:
            final_response = {"isento_icms": "Não", "convenio": None}
        else:
            final_response = {"isento_icms": "Erro Inesperado", "convenio": None}

        return final_response

    except Exception as e:
        print(f"Erro geral ao processar produto com o agente: {e}")
        return {"isento_icms": "Erro Inesperado", "convenio": None}
    

def consulta_produtos_excel():
    caminho_arquivo = r"\\10.1.1.202\c\rpa\contabilidade\Cadastro Produtos\Base Para ajuste.xlsx"
    # caminho_arquivo = r"C:\Users\Nícolas Nasário\Downloads\Base cadastros - (NOVOS MEDICAMENTO) 22.10.2025.xlsx"
    
    workbook = load_workbook(caminho_arquivo)
    sheet = workbook["BASE CADASTRO"]

    linha_inicial = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        codigo_produto = row[2]
        digito_produto = row[4]
        nome_produto = row[5]
        apresentacao = row[6]
        ncm = row[10]
        principio_ativo = row[13]

        codigo_completo = f"{codigo_produto}{digito_produto}"
        nome_completo = f"{nome_produto}{apresentacao}"

        produto = {
            "codigo": f"{codigo_completo}",
            "nome": f"{nome_completo}",
            "principio_ativo": f"{principio_ativo}",
            "ncm": f"{ncm}"
        }

        resultado = process_product_with_rag(**produto) # Não precisa passar agent_executor aqui
        print(f"Produto: {produto['codigo']} - Isento: {resultado['isento_icms']} - Convênio: {resultado['convenio']}")

        if resultado['isento_icms'] != "Erro de Formato":
            resultado = f"""{resultado['isento_icms']}.{resultado['convenio'] if resultado['convenio'] else ''}"""

        elif resultado['isento_icms'] == "Erro de Formato":
            print(f"Erro de Formato para o produto {produto['codigo']}. Resultado bruto: {resultado}")
            resultado = process_product_with_rag(**produto) # Não precisa passar agent_executor aqui
            print(f"Produto: {produto['codigo']} - Isento: {resultado['isento_icms']} - Convênio: {resultado['convenio']}")

            if resultado['isento_icms'] != "Erro de Formato":
                resultado = f"""{resultado['isento_icms']}.{resultado['convenio'] if resultado['convenio'] else ''}"""

        
        if "sim" in str(resultado).lower():
            if len(str(resultado)) < 5:
                resultado = "Não"

        sheet.cell(row=linha_inicial, column=35).value = str(resultado)
        # sheet.cell(row=linha_inicial, column=41).value = str(resultado)

        linha_inicial += 1
    
    workbook.save(caminho_arquivo)
    workbook.close()


def move_arquivo_manda_email():
    anexo = []
    path = r"\\10.1.1.202\c\rpa\contabilidade\Cadastro Produtos"
    lista_arquivos = os.listdir(path)

    for arquivo in lista_arquivos:
        if arquivo.endswith(".xlsx"):
            caminho_arquivo = os.path.join(path, arquivo)
            anexo.append(caminho_arquivo)

            break

    assunto = f"Cadastro de produtos fiscal"
    mensagem = f"""
    Olá! <br><br>

    Segue em anexo o arquivo para cadastro das mercadorias que estão pendentes no financeiro. <br><br>
    """

    destinatarios = []
    # destinatarios.append("nicolas.nasario@COMPANY_NAME.com.br")
    destinatarios.append("israel.martins@COMPANY_NAME.com.br")
    destinatarios.append("kaue.baesso@COMPANY_NAME.com.br")
    destinatarios.append("joao.bernardo@COMPANY_NAME.com.br")
    destinatarios.append("davi.lopes@COMPANY_NAME.com.br")
    destinatarios.append("fiscal@COMPANY_NAME.com.br")

    zimbra = ZimbraMailer()
    zimbra.envia_email(assunto_email=assunto, mensagem_email=mensagem, destinatarios_email=destinatarios, anexos=anexo)

    data_atual = datetime.now()
    data_convertida = data_atual.strftime("%d_%m_%Y_%H_%M_%S")
    local_destino = rf"\\10.1.1.202\c\rpa\contabilidade\Cadastro Produtos\Backup\{data_convertida}"

    if not os.path.exists(local_destino):
        os.makedirs(local_destino)

    mover = HubArquivos(caminho_destino=local_destino, arquivo=arquivo, caminho_geral=path)
    mover.verificarDiretorio()


def main():
    initialize_rag_components() # Inicializa os componentes RAG
    consulta_produtos_excel()
    move_arquivo_manda_email()


# Para testar rag_core.py individualmente (opcional):
if __name__ == "__main__":
    # Inicializa os componentes APENAS se este script for executado diretamente
    print("Executando rag_core.py diretamente para teste...")
    main()
    
    # test_product = {
    #     "codigo": "765951",
    #     "nome": "KOSELUGO10MG 60CAPS",
    #     "principio_ativo": "SULFATO DE SELUMETINIBE 10 MG",
    #     "ncm": "30049079"
    # }
    # test_product = {
    #     "codigo": "774112",
    #     "nome": "AMOXI.+CLAV.MEDLEY",
    #     "principio_ativo": "AMOXICILINA TRI-HIDRATADA 80 MG/ML + CLAVULANATO DE POTASSIO 11.4 MG/ML",
    #     "ncm": "30041012"
    # }
    # result = process_product_with_rag(**test_product) # Não precisa passar agent_executor aqui
    # print("\nResultado do teste direto (rag_core.py):")
    # print(json.dumps(result, indent=2))